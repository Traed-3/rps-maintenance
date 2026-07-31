#!/usr/bin/env python3
"""
relink-invoice-customers — take the customer from each invoice's own sheet.

Two mistakes in the first pass:

  1. The Dropbox folder was treated as the stronger signal. It isn't.
     "4641 Invoice.xls" sits in SEI In House Maintenance but its CUSTOMER
     cell says Tanknology, and the invoice is right — RPS did the work for
     Tanknology at a 7-Eleven site.

  2. Customers were resolved per JOB, using whichever invoice file turned
     up first, then applied to every invoice on that job. One file naming
     a sub spread that name across five unrelated invoices.

So: match each invoice to the sheet whose grand total equals it, read the
CUSTOMER cell from that sheet, and set the customer on that invoice alone.
The folder is only consulted when the sheet has no readable customer.

Dry run unless --commit.

  python3 scripts/ingest/relink-invoice-customers.py [--commit]
"""
import json, os, re, sys, urllib.request, warnings

warnings.filterwarnings('ignore')
DRY = '--commit' not in sys.argv

ENV = {}
for line in open(os.path.expanduser('~/Developer/rps-maintenance/.env.local'), encoding='utf8'):
    line = line.strip()
    if line and not line.startswith('#') and '=' in line:
        k, v = line.split('=', 1)
        ENV[k] = v.strip().strip('"')
URL, KEY = ENV['NEXT_PUBLIC_SUPABASE_URL'], ENV['SUPABASE_SERVICE_ROLE_KEY']
H = {'apikey': KEY, 'Authorization': f'Bearer {KEY}', 'Content-Type': 'application/json'}


def api(path, method='GET', body=None, prefer=None):
    req = urllib.request.Request(f'{URL}/rest/v1/{path}', method=method,
                                 data=json.dumps(body).encode() if body is not None else None)
    for k, v in H.items():
        req.add_header(k, v)
    if prefer:
        req.add_header('Prefer', prefer)
    with urllib.request.urlopen(req) as r:
        raw = r.read().decode()
        return json.loads(raw) if raw.strip() else None


def page(table, select):
    out, frm = [], 0
    while True:
        req = urllib.request.Request(f'{URL}/rest/v1/{table}?select={select}')
        for k, v in H.items():
            req.add_header(k, v)
        req.add_header('Range', f'{frm}-{frm+999}')
        with urllib.request.urlopen(req) as r:
            b = json.loads(r.read().decode())
        out += b
        if len(b) < 1000:
            return out
        frm += 1000


def open_sheet(path):
    if path.lower().endswith('.xls'):
        import xlrd
        s = xlrd.open_workbook(path).sheet_by_index(0)
        return s.nrows, s.ncols, (lambda r, c: s.cell_value(r, c) if r < s.nrows and c < s.ncols else '')
    import openpyxl
    s = openpyxl.load_workbook(path, data_only=True).worksheets[0]
    R, C = min(s.max_row or 0, 200), min(s.max_column or 0, 30)
    return R, C, (lambda r, c: (s.cell(row=r + 1, column=c + 1).value if r < R and c < C else '') or '')


def num(v):
    return None if isinstance(v, bool) else (v if isinstance(v, (int, float)) else None)


TOTALS = ('PROPOSAL GRAND TOTAL', 'INVOICE GRAND TOTAL', 'GRAND TOTAL', 'INVOICE TOTAL', 'TOTAL DUE')


def grand_total(get, R, C):
    for want in TOTALS:
        for r in range(R):
            for c in range(C):
                v = get(r, c)
                if isinstance(v, str) and want in v.strip().upper():
                    for cc in range(c + 1, C):
                        n = num(get(r, cc))
                        if n:
                            return n
    return None


def customer_cell(get, R, C):
    """The value directly under the CUSTOMER: label, else to its right."""
    for r in range(R):
        for c in range(C):
            v = get(r, c)
            if isinstance(v, str) and v.strip().upper().rstrip(': ').startswith('CUSTOMER'):
                for rr in range(r + 1, min(r + 3, R)):
                    t = get(rr, c)
                    if isinstance(t, str) and t.strip() and len(t.strip()) < 45:
                        return t.strip()
                for cc in range(c + 1, C):
                    t = get(r, cc)
                    if isinstance(t, str) and t.strip() and len(t.strip()) < 45:
                        return t.strip()
    return None


# Spelling variants -> the customer we file under. Anything not listed here is
# taken at face value and becomes its own customer: the invoice is the record.
ALIASES = [
    (r'sei\s*[-–]?\s*project\s*forum',      '7-Eleven', 'SEI Project Forum'),
    (r'sei\s*[-–]?\s*in\s*[-–]?\s*house',   '7-Eleven', 'SEI In-House'),
    (r'sei\s*[-–]?\s*compliance',           '7-Eleven', 'SEI Compliance'),
    (r'\bsei\b|^project\s*forum$',          '7-Eleven', 'SEI'),
    (r'7\s*[-–]?\s*eleven.*environmental',  '7-Eleven', 'Environmental'),
    (r'7\s*[-–]?\s*eleven|7\s*eleven',      '7-Eleven', None),
    (r'\bvixxo\b',                          'Vixxo', None),
    (r'\bsheetz\b',                         'Sheetz', None),
    (r'\bwawa\b',                           'Wawa', None),
    (r'\bsunoco\b',                         'Sunoco LP', None),
    (r'capit[oa]l\s+petroleum',             'Capital Petroleum Group', None),
    (r'global\s+partners?',                 'Global Partners', None),
    (r'\btanknology\b',                     'Tanknology', None),
    # Vericon Construction were the GC on the 17703 ground-up and hired RPS for
    # part of their scope, so they are a customer in their own right. Two
    # spellings appear across their invoices.
    (r'\bvericon\b',                        'Vericon Construction', None),
    (r'\bics\b|imagine\s+commer',           'ICS - Imagine Commercial Solutions', None),
    (r'\baecom\b',                          'AECOM', None),
    (r'jf\s+petroleum',                     'JF Petroleum Group', None),
    (r'jefferson\s+county',                 'Jefferson County Schools', None),
    (r'\bspeedway\b',                       'IDS-Speedway', None),
    (r'travel\s*america',                   'Travel America', None),
]
# Cells that are plainly not a customer name.
JUNK = re.compile(r'^(project manager|site\s*#|location:?|customer:?|attn:?|n/?a|store|'
                  r'[\d\-()\s]+|[A-Z]{2,}\d{5,})$', re.I)


def resolve(raw):
    if not raw or JUNK.match(raw.strip()):
        return None, None
    t = raw.strip()
    for pat, name, prog in ALIASES:
        if re.search(pat, t, re.I):
            return name, prog
    return re.sub(r'\s+', ' ', t).strip(' .,'), None


print(f"\n== relink-invoice-customers {'(DRY RUN)' if DRY else '(COMMIT)'} ==")

customers = {c['name'].lower(): c for c in page('con_customers', 'id,name')}
invoices = page('con_invoices', 'id,job_id,customer_id,invoice_grand_total')
docs = {}
for d in page('con_documents', 'id,job_id,file_name,source_path,category'):
    if d['category'] == 'invoices' and d.get('source_path') and re.search(r'\.xlsx?$', d['file_name'], re.I):
        docs.setdefault(d['job_id'], []).append(d)
COMPANY = page('con_jobs', 'company_id')[0]['company_id']
cur_name = {c['id']: c['name'] for c in customers.values()}

job_site = {j['id']: j['site_number'] for j in page('con_jobs', 'id,site_number')}
plan, nosheet, nocell = [], 0, 0
overridden_4xxxx, overridden_owner = [], []
for inv in invoices:
    want = float(inv['invoice_grand_total'] or 0)
    raw = None
    for d in docs.get(inv['job_id'], []):
        if not os.path.exists(d['source_path']):
            continue
        try:
            R, C, get = open_sheet(d['source_path'])
            if abs((grand_total(get, R, C) or -1) - want) > 0.01:
                continue
            raw = customer_cell(get, R, C)
        except Exception:
            continue
        if raw:
            break
    if raw is None:
        nosheet += 1
        continue
    name, prog = resolve(raw)
    if not name:
        nocell += 1
        continue

    # Two rulings from Trae that override what the cell says.
    #  - Every 4xxxx site is Sunoco. Their invoices often say SEI because the
    #    two were billed together back when 7-Eleven had just bought them.
    #  - The Sheetz 210 invoice names the property owner, not the payer.
    site = str(job_site.get(inv['job_id']) or '')
    m = re.search(r'(?<!\d)(\d{5})(?!\d)', site)
    if m and m.group(1).startswith('4'):
        overridden_4xxxx.append(inv)
        continue
    if re.search(r'wu\s+ching', name, re.I):
        overridden_owner.append(inv)
        continue

    plan.append((inv, name, prog, raw))

changes = [p for p in plan if cur_name.get(p[0]['customer_id']) != p[1]]
need = sorted({n for _, n, _, _ in plan if n.lower() not in customers})
print(f'invoices read from their own sheet : {len(plan)}')
print(f'  customer would CHANGE            : {len(changes)}')
print(f'no sheet matched the total         : {nosheet}')
print(f'cell unusable                      : {nocell}')
print(f'kept Sunoco (4xxxx rule)           : {len(overridden_4xxxx)}')
print(f'kept Sheetz (owner name in cell)   : {len(overridden_owner)}')
if need:
    print(f'\ncustomers to create: {", ".join(need)}')

moves = {}
for inv, name, _, raw in changes:
    k = f'{cur_name.get(inv["customer_id"]) or "(none)"} -> {name}'
    moves[k] = moves.get(k, 0) + 1
print('\nchanges:')
for k, v in sorted(moves.items(), key=lambda x: -x[1])[:18]:
    print(f'   {str(v).rjust(4)}  {k}')

if DRY:
    print('\nDRY RUN — nothing written. Re-run with --commit')
    sys.exit(0)

for name in need:
    made = api('con_customers', 'POST', {'company_id': COMPANY, 'name': name}, prefer='return=representation')
    customers[name.lower()] = made[0]
    print(f'  created customer: {name}')

done = 0
for inv, name, prog, _ in changes:
    cid = customers[name.lower()]['id']
    api(f"con_invoices?id=eq.{inv['id']}", 'PATCH', {'customer_id': cid})
    done += 1
print(f'\nrelinked {done} invoices')
