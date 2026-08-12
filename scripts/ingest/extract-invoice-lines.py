#!/usr/bin/env python3
"""
extract-invoice-lines — fill in each invoice line by line.

The first pass stored only the grand total, so an invoice showed $4,650
and nothing about what made it up. The "Capital Proposal" sheet lays the
detail out in three blocks:

    REMOVAL AND DISPOSAL     ┐ both stored as section 'basic',
    BASIC INSTALLATION       ┘ told apart by item_type
    ADDITIONAL SCOPE OF WORK → section 'additional'

with the columns the app already has: description, quantity, unit cost,
material, labor hours, labor rate, total labor, total material+labor.

Only touches invoices that have no lines yet, so it is safe to re-run.
Dry run unless --commit.

  python3 scripts/ingest/extract-invoice-lines.py [--commit] [--limit N]
"""
import json, os, re, sys, urllib.request, warnings

warnings.filterwarnings('ignore')
DRY = '--commit' not in sys.argv
LIMIT = int(sys.argv[sys.argv.index('--limit') + 1]) if '--limit' in sys.argv else None

ENV = {}
for line in open(os.path.expanduser('~/Developer/rps-maintenance/.env.local'), encoding='utf8'):
    line = line.strip()
    if line and not line.startswith('#') and '=' in line:
        k, v = line.split('=', 1)
        ENV[k] = v.strip().strip('"')
URL, KEY = ENV['NEXT_PUBLIC_SUPABASE_URL'], ENV['SUPABASE_SERVICE_ROLE_KEY']
H = {'apikey': KEY, 'Authorization': f'Bearer {KEY}', 'Content-Type': 'application/json'}


def api(path, method='GET', body=None):
    req = urllib.request.Request(f'{URL}/rest/v1/{path}', method=method,
                                 data=json.dumps(body).encode() if body is not None else None)
    for k, v in H.items():
        req.add_header(k, v)
    with urllib.request.urlopen(req) as r:
        raw = r.read().decode()
        return json.loads(raw) if raw.strip() else None


def page(table, select):
    out, frm = [], 0
    while True:
        req = urllib.request.Request(f'{URL}/rest/v1/{table}?select={select}')
        for k, v in H.items():
            req.add_header(k, v)
        req.add_header('Range', f'{frm}-{frm+999}')
        with urllib.request.urlopen(req) as r:
            b = json.loads(r.read().decode())
        out += b
        if len(b) < 1000:
            return out
        frm += 1000


def open_sheet(path):
    if path.lower().endswith('.xls'):
        import xlrd
        s = xlrd.open_workbook(path).sheet_by_index(0)
        return s.nrows, s.ncols, (lambda r, c: s.cell_value(r, c) if r < s.nrows and c < s.ncols else '')
    import openpyxl
    s = openpyxl.load_workbook(path, data_only=True).worksheets[0]
    R, C = min(s.max_row or 0, 200), min(s.max_column or 0, 30)
    return R, C, (lambda r, c: (s.cell(row=r + 1, column=c + 1).value if r < R and c < C else '') or '')


def num(v):
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return v
    return None


# Where each block starts, and what it means to the app.
BLOCKS = [
    (re.compile(r'REMOVAL\s+AND\s+DISPOSAL(?!\s+TOTAL)', re.I), 'basic',      'removal'),
    (re.compile(r'BASIC\s+INSTALLATION', re.I),                 'basic',      'basic'),
    (re.compile(r'ADDITIONAL\s+SCOPE\s+OF\s+WORK(?!\s+TOTAL)', re.I), 'additional', 'additional'),
]
STOP = re.compile(r'SUB\s*TOTAL|TOTAL:|PROFIT AND OVERHEAD|SALES TAX|GRAND TOTAL', re.I)

# Column letters in the RPS template: C description, I qty, J unit cost,
# K material, L labor hours, M labor rate, N total labor, O total M+L.
COL = {'desc': 2, 'qty': 8, 'unit': 9, 'material': 10, 'hours': 11, 'rate': 12, 'labor': 13, 'total': 14}


def lines_in(path):
    """Single pass down the sheet.

    Each section header switches the block we are in; rows belong to whichever
    block was declared most recently. Scanning a fixed window forward from each
    header instead would double-count every row that two windows overlap — which
    is exactly what produced 200% totals.
    """
    R, C, get = open_sheet(path)
    if C <= COL['total']:
        return []          # not the standard template — skip rather than guess

    out = []
    section = kind = None
    for r in range(R):
        # Scan the WHOLE row: "Sub Total Material:" sits in column H, well past
        # the first few columns, and missing it lets a section's own total be
        # read as another line — doubling that section.
        head = ' '.join(str(get(r, c)) for c in range(C) if isinstance(get(r, c), str))

        matched = None
        for rx, sec, k in BLOCKS:
            if rx.search(head):
                matched = (sec, k)
                break
        if matched:
            section, kind = matched
            continue                        # the header row itself is not a line

        if section is None:
            continue
        if STOP.search(head):               # subtotal / tax / grand total ends the block
            section = kind = None
            continue

        desc = get(r, COL['desc'])
        desc = desc.strip() if isinstance(desc, str) else ''
        total = num(get(r, COL['total'])) or 0
        material = num(get(r, COL['material'])) or 0
        labor = num(get(r, COL['labor'])) or 0
        # A real line always names what it is. An unlabelled row carrying a
        # figure is the section total, not an item.
        if not desc:
            continue
        out.append({
            'section': section,
            'line_no': int(num(get(r, 1)) or 0) or None,
            'description': desc or None,
            'quantity': num(get(r, COL['qty'])),
            'unit_cost': num(get(r, COL['unit'])),
            'material_total': material,
            'labor_hours': num(get(r, COL['hours'])),
            'labor_rate': num(get(r, COL['rate'])),
            'total_labor': labor,
            'total_material_labor': total,
            'item_type': kind,
        })
    return out


print(f"\n== extract-invoice-lines {'(DRY RUN)' if DRY else '(COMMIT)'} ==")

invoices = page('con_invoices', 'id,job_id,invoice_grand_total')
have_lines = {l['invoice_id'] for l in page('con_invoice_line_items', 'invoice_id')}
docs = {}
for d in page('con_documents', 'id,job_id,file_name,source_path,category'):
    if d['category'] == 'invoices' and d.get('source_path') and re.search(r'\.xlsx?$', d['file_name'], re.I):
        docs.setdefault(d['job_id'], []).append(d)

todo = [i for i in invoices if i['id'] not in have_lines and i['job_id'] in docs]
if LIMIT:
    todo = todo[:LIMIT]
print(f'invoices without lines, with a spreadsheet: {len(todo)}')

TOTAL_LABELS = ('PROPOSAL GRAND TOTAL', 'INVOICE GRAND TOTAL', 'GRAND TOTAL',
                'INVOICE TOTAL', 'TOTAL DUE')


def grand_total_of(path):
    """The sheet's own grand total, used to prove which invoice it is."""
    R, C, get = open_sheet(path)
    for want in TOTAL_LABELS:
        for r in range(R):
            for c in range(C):
                v = get(r, c)
                if isinstance(v, str) and want in v.strip().upper():
                    for cc in range(c + 1, C):
                        n = num(get(r, cc))
                        if n:
                            return n
    return None


plan, nolines, unmatched_total, failed = [], 0, 0, 0
for inv in todo:
    want = float(inv['invoice_grand_total'] or 0)
    got, source = [], None

    # A job can hold several invoices — an original plus change orders. Only
    # take lines from the sheet whose own grand total IS this invoice's, or
    # the detail lands under the wrong invoice entirely.
    for d in docs[inv['job_id']]:
        if not os.path.exists(d['source_path']):
            continue
        try:
            if abs((grand_total_of(d['source_path']) or -1) - want) > 0.01:
                continue
            got = lines_in(d['source_path'])
            source = d
        except Exception:
            failed += 1
            continue
        if got:
            break

    if source is None:
        unmatched_total += 1
        continue
    if not got:
        nolines += 1
        continue
    plan.append((inv, got))

total_lines = sum(len(l) for _, l in plan)
print(f'  invoices that yielded lines : {len(plan)}')
print(f'  line items to create        : {total_lines}')
print(f'  no lines found              : {nolines}')
print(f'  no sheet matched the total  : {unmatched_total}')
print(f'  unreadable                  : {failed}')

if DRY:
    print('\nsample — lines for one invoice:')
    if plan:
        inv, lines = max(plan, key=lambda x: len(x[1]))
        print(f"   invoice total ${inv['invoice_grand_total']}")
        for l in lines[:10]:
            print(f"     [{l['section']}/{l['item_type']}] {(l['description'] or '')[:38]:<40} "
                  f"mat {l['material_total']:>9} labor {l['total_labor']:>8} = {l['total_material_labor']}")
    print('\nDRY RUN — nothing written. Re-run with --commit')
    sys.exit(0)

made = 0
for inv, lines in plan:
    body = [{**l, 'invoice_id': inv['id']} for l in lines]
    try:
        api('con_invoice_line_items', 'POST', body)
        made += len(body)
        if made % 500 < len(body):
            print(f'  {made}/{total_lines} lines')
    except Exception as e:
        failed += 1

print(f'\ncreated {made} line items across {len(plan)} invoices')
