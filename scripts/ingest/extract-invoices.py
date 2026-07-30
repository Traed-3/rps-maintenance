#!/usr/bin/env python3
"""
extract-invoices — turn imported invoice spreadsheets into real invoice records.

The importer filed 720 invoice files as documents, so the app knew the
filename and nothing else. Every historical total sat locked inside a
spreadsheet. This reads them and creates con_invoices rows against the job
the document is already attached to.

RPS has three template families in the wild:
  "Capital Proposal"        -> PROPOSAL GRAND TOTAL:
  "INVOICE"                 -> INVOICE TOTAL:
  Sheet1 / Auto Invoice ... -> TOTAL: / SUBTOTAL

Reads only. Dry run unless --commit is passed.

  python3 scripts/ingest/extract-invoices.py [--commit] [--limit N]
"""
import json, os, re, sys, urllib.request, datetime

DRY = '--commit' not in sys.argv
LIMIT = None
if '--limit' in sys.argv:
    LIMIT = int(sys.argv[sys.argv.index('--limit') + 1])

# ── env ─────────────────────────────────────────────────────
ENV = {}
for line in open(os.path.expanduser('~/Developer/rps-maintenance/.env.local'), encoding='utf8'):
    line = line.strip()
    if line and not line.startswith('#') and '=' in line:
        k, v = line.split('=', 1)
        ENV[k] = v.strip().strip('"')
URL = ENV['NEXT_PUBLIC_SUPABASE_URL']
KEY = ENV['SUPABASE_SERVICE_ROLE_KEY']
H = {'apikey': KEY, 'Authorization': f'Bearer {KEY}', 'Content-Type': 'application/json'}


def api(path, method='GET', body=None, extra=None):
    req = urllib.request.Request(f'{URL}/rest/v1/{path}', method=method,
                                 data=json.dumps(body).encode() if body is not None else None)
    for k, v in {**H, **(extra or {})}.items():
        req.add_header(k, v)
    with urllib.request.urlopen(req) as r:
        raw = r.read().decode()
        return json.loads(raw) if raw.strip() else None


def page(table, select):
    out, frm = [], 0
    while True:
        req = urllib.request.Request(f'{URL}/rest/v1/{table}?select={select}')
        for k, v in H.items():
            req.add_header(k, v)
        req.add_header('Range', f'{frm}-{frm+999}')
        with urllib.request.urlopen(req) as r:
            b = json.loads(r.read().decode())
        out += b
        if len(b) < 1000:
            return out
        frm += 1000


# ── spreadsheet access, one interface for xls and xlsx ──────
def open_sheet(path):
    if path.lower().endswith('.xls'):
        import xlrd
        s = xlrd.open_workbook(path).sheet_by_index(0)
        return (s.name, s.nrows, s.ncols,
                lambda r, c: s.cell_value(r, c) if r < s.nrows and c < s.ncols else '')
    import openpyxl, warnings
    warnings.filterwarnings('ignore')
    s = openpyxl.load_workbook(path, data_only=True).worksheets[0]
    R, C = min(s.max_row or 0, 200), min(s.max_column or 0, 30)
    return (s.title, R, C, lambda r, c: (s.cell(row=r + 1, column=c + 1).value
                                         if r < R and c < C else '') or '')


def num(v):
    return v if isinstance(v, (int, float)) and not isinstance(v, bool) else None


def excel_date(v):
    """Excel serial -> ISO date. Ignores anything outside a sane range."""
    if isinstance(v, datetime.datetime):
        return v.date().isoformat()
    if isinstance(v, datetime.date):
        return v.isoformat()
    n = num(v)
    if n is None or not (20000 < n < 60000):
        return None
    return (datetime.date(1899, 12, 30) + datetime.timedelta(days=int(n))).isoformat()


# Total labels in priority order — most specific first, so a sheet carrying
# both "SUB TOTAL" and "PROPOSAL GRAND TOTAL" resolves to the real total.
TOTAL_LABELS = [
    'PROPOSAL GRAND TOTAL', 'INVOICE GRAND TOTAL', 'GRAND TOTAL',
    'INVOICE TOTAL', 'TOTAL DUE', 'AMOUNT DUE',
]


def find_labeled_number(get, R, C, labels, exact=False):
    """Value to the right of (or below) the first matching label."""
    for want in labels:
        for r in range(R):
            for c in range(C):
                v = get(r, c)
                if not isinstance(v, str):
                    continue
                t = v.strip().upper().rstrip(':')
                hit = (t == want) if exact else (want in t)
                if not hit:
                    continue
                for cc in range(c + 1, C):          # to the right
                    n = num(get(r, cc))
                    if n:
                        return n, v.strip()
                for rr in range(r + 1, min(r + 3, R)):   # or just below
                    n = num(get(rr, c))
                    if n:
                        return n, v.strip()
    return None, None


def find_label_text(get, R, C, labels):
    for want in labels:
        for r in range(R):
            for c in range(C):
                v = get(r, c)
                if isinstance(v, str) and want in v.strip().upper():
                    for cc in range(c + 1, C):
                        t = get(r, cc)
                        if isinstance(t, str) and t.strip():
                            return t.strip()
    return None


def find_date(get, R, C):
    for want in ('INVOICE DATE', 'DATE:'):
        for r in range(R):
            for c in range(C):
                v = get(r, c)
                if isinstance(v, str) and want in v.strip().upper():
                    for cc in range(c + 1, C):
                        d = excel_date(get(r, cc))
                        if d:
                            return d
    return None


def extract(path):
    name, R, C, get = open_sheet(path)
    total, label = find_labeled_number(get, R, C, TOTAL_LABELS)
    if total is None:                     # fall back to a bare TOTAL row
        total, label = find_labeled_number(get, R, C, ['TOTAL'], exact=True)
    return {
        'sheet': name,
        'total': total,
        'total_label': label,
        'invoice_date': find_date(get, R, C),
        'csr_number': find_label_text(get, R, C, ['SR #', 'CSR', 'PO #']),
        'store_label': find_label_text(get, R, C, ['STORE']),
        'facility_address': find_label_text(get, R, C, ['ADDRESS']),
        'city_state_zip': find_label_text(get, R, C, ['CITY,ST,ZIP', 'CITY, ST, ZIP']),
        'attn': find_label_text(get, R, C, ['ATTN']),
        'project_description': find_label_text(get, R, C, ['PROJECT DESCRIPTION']),
        'profit_overhead_percent': (find_labeled_number(get, R, C, ['PROFIT AND OVERHEAD PERCENT'])[0] or 0),
        'sales_tax_percent': (find_labeled_number(get, R, C, ['SALES TAX PERCENT'])[0] or 0),
    }


# ── main ────────────────────────────────────────────────────
print(f"\n== extract-invoices {'(DRY RUN)' if DRY else '(COMMIT)'} ==")

COMPANY = page('con_jobs', 'company_id')[0]['company_id']

# An autosave or a Dropbox conflicted copy is the SAME invoice under another
# name. Entering both would double-count revenue in the costing figures.
DUPLICATE_NAME = re.compile(r'autosaved|conflicted copy', re.I)
TEMPLATE_NAME = re.compile(r'\btemplate\b|\bsample\b|\bexample\b|\bblank\b', re.I)

docs = [d for d in page('con_documents', 'id,job_id,file_name,source_path,category')
        if d['category'] == 'invoices'
        and d.get('source_path')
        and re.search(r'\.xlsx?$', d['file_name'], re.I)]

dupes = [d for d in docs if DUPLICATE_NAME.search(d['file_name'])]
docs = [d for d in docs if not DUPLICATE_NAME.search(d['file_name'])]

if LIMIT:
    docs = docs[:LIMIT]
print(f'invoice spreadsheets to read: {len(docs)}')
if dupes:
    print(f'skipping {len(dupes)} autosave/conflicted copies (same invoice, different name):')
    for d in dupes[:5]:
        print(f'    {d["file_name"][:70]}')

made = skipped = nototal = missing = failed = 0
rows = []
for d in docs:
    p = d['source_path']
    if not os.path.exists(p):
        missing += 1
        continue
    try:
        info = extract(p)
    except Exception as e:
        failed += 1
        continue
    if not info['total']:
        nototal += 1
        continue
    if not d['job_id']:
        skipped += 1
        continue
    rows.append((d, info))

print(f'  parsed with a total : {len(rows)}')
print(f'  no total found      : {nototal}')
print(f'  file missing on disk: {missing}')
print(f'  failed to open      : {failed}')
print(f'  no job attached     : {skipped}')

if DRY:
    print('\nsample of what would be created:')
    for d, i in rows[:12]:
        print(f"   {str(i['total']):>12}  {i['invoice_date'] or '(no date)':>10}  "
              f"{(i['store_label'] or '')[:10]:<10} {d['file_name'][:44]}")
    tot = sum(i['total'] for _, i in rows)
    print(f'\n   total invoiced value across {len(rows)} files: ${tot:,.2f}')
    print('\nDRY RUN — nothing written. Re-run with --commit')
    sys.exit(0)

for d, i in rows:
    body = {
        'company_id': COMPANY,
        'job_id': d['job_id'],
        'invoice_date': i['invoice_date'],
        'csr_number': i['csr_number'],
        'store_label': i['store_label'],
        'facility_address': i['facility_address'],
        'city_state_zip': i['city_state_zip'],
        'attn': i['attn'],
        'project_description': i['project_description'],
        'profit_overhead_percent': i['profit_overhead_percent'],
        'sales_tax_percent': i['sales_tax_percent'],
        'invoice_grand_total': round(i['total'], 2),
        'grand_total': round(i['total'], 2),
        'status': 'sent',
    }
    try:
        api('con_invoices', 'POST', body)
        made += 1
        if made % 25 == 0:
            print(f'  created {made}/{len(rows)}')
    except Exception as e:
        failed += 1

print(f'\ncreated {made} invoice records')
